import time
import openpyxl
from datetime import datetime
import os
import pandas as pd





dataframe = pd.DataFrame(
        columns=['timestamp', 'carbon_intensity', 'low_emissions', 'renewable_emissions', 'total_production',
                 'total_emissions'
            , 'nucleare_installed_capacity', 'nucleare_production', 'nucleare_emissions',
                 'geotermico_installed_capacity', 'geotermico_production', 'geotermico_emissions'
            , 'biomassa_installed_capacity', 'biomassa_production', 'biomassa_emissions', 'carbone_installed_capacity',
                 'carbone_production', 'carbone_emissions'
            , 'eolico_installed_capacity', 'eolico_production', 'eolico_emissions', 'fotovoltaico_installed_capacity',
                 'fotovoltaico_production', 'fotovoltaico_emissions'
            , 'idroelettrico_installed_capacity', 'idroelettrico_production', 'idroelettrico_emissions',
                 'accumuloidro_installed_capacity', 'accumuloidro_production', 'accumuloidro_emissions'
            , 'batterieaccu_installed_capacity', 'batterieaccu_production', 'batterieaccu_emissions',
                 'gas_installed_capacity', 'gas_production', 'gas_emissions'
            , 'petrolio_installed_capacity', 'petrolio_production', 'petrolio_emissions',
                 'sconosciuto_installed_capacity', 'sconosciuto_production', 'sconosciuto_emissions'
            , 'exchange_export', 'exchange_import'])

stati = ['El Hierro (Spagna)', 'Svezia', 'Francia','Danimarca orientale (Danimarca)', 'Romania',
             'Spagna', 'Belgio', 'Lettonia', 'Portogallo', 'Gran Canaria (Spagna)',
             'Danimarca occidentale (Danimarca)', 'Ungheria', 'Formentera (Spagna)', 'Lussemburgo', 'Finlandia',
             'Austria', 'Slovenia', 'Maiorca (Spagna)', 'Paesi Bassi', 'Lituania',
             'Slovacchia', 'Croazia', 'Ibiza (Spagna)', 'Polonia', 'Irlanda',
             'Sicilia (Italia)', 'Bulgaria', 'Tenerife (Spagna)', 'Germania', 'Centronord (Italia)',
             'Cipro', 'Minorca (Spagna)', 'Settentrione (Italia)', 'Estonia', 'Fuerteventura Lanzarote (Spagna)',
             'La Palma (Spagna)', 'Meridione (Italia)', 'La Gomera (Spagna)', 'Centrosud (Italia)', 'Sardegna (Italia)']

STATES_DIR = "../statesRiparati"

if __name__ == '__main__':
    #statii = ['Germania']
    start = time.time()
    for s in stati:
        try:
            print(s)
            file_excel = openpyxl.load_workbook('states/'+s+'.xlsx')
            sheet_obj = file_excel.active
            numr=sheet_obj.max_row+1
            numc=sheet_obj.max_column+1


            # ripara timestamp non alliniati a 00 10 20 30 40 50 + gestisce l'ultima colonna  total_production se contiene i
            # valori al posto di total_capacity (errore previsto per via aggiornamento codice non necessario in caso di ripartenza da 0 con i file excel)
            rnew=[]

            c = []
            for j in range(1, numc-1):
                    if(j!=5):
                        c.append(sheet_obj.cell(row=1, column=j).value)
                    else:
                        c.append(sheet_obj.cell(row=1, column=numc-1).value)
            rnew.append(c)


            for i in range(2,numr):
                c=[]
                for j in range(1, numc-1):
                    tmp=sheet_obj.cell(row=i, column=j).value

                    if j != 1 and j != 5:
                        c.append(tmp)
                    else:
                        if(j==1 and tmp[4] != "0"):
                            c.append(tmp[0:4] + "0" + tmp[5:])
                        else:
                            if(j==5 and tmp==None):
                                c.append(sheet_obj.cell(row=i, column=numc - 1).value)
                            else:
                                c.append(tmp)
                rnew.append(c)

            file_excel.close()
            #######################################################################################################################################################################

            # riparare i buchi causati dal crash del sistema
            starttime=timestamp=int(round(datetime.strptime(rnew[1][0],'%H:%M %d-%m-%Y').timestamp()))

            nuovofile=[]
            nuovofile.append(rnew[0])

            for i in range(1,numr-1):
                timestamp=int(round(datetime.strptime(rnew[i][0],'%H:%M %d-%m-%Y').timestamp()))
                starttime += 600

                if(timestamp==starttime):
                    nuovofile.append(rnew[i])

                else:
                    while(timestamp>starttime):
                        timestampnew = datetime.fromtimestamp(starttime).strftime('%H:%M %d-%m-%Y')
                        nuovofile.append([timestampnew, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])

                        starttime += 600
                    nuovofile.append(rnew[i])
            #######################################################################################################################################################################

            #qui sotto gestisco il salvataggio del nuovo file
            #s=s+"NEW"

            path = os.path.join(STATES_DIR, s + ".xlsx")

            dfaaa = pd.DataFrame(columns=dataframe.columns)#44 colonne
            for i in range(1, len(nuovofile)):
                j = 0
                tmp = {}
                for df in dataframe:
                    tmp[df] = nuovofile[i][j]
                    j += 1
                dfbbb = pd.concat([dataframe.copy(), pd.DataFrame(tmp, index=[0])], ignore_index=True)
                dfaaa = pd.concat([dfaaa,  dfbbb])

            try:
                if (len(s) <= 30):
                    dfaaa.to_excel(path, sheet_name=s, index=False)
                else:
                    s1 = s.split(" ")[0]
                    dfaaa.to_excel(path, sheet_name=s1, index=False)

            except Exception as e:
                pass
        except Exception as e:
            print(e)
    end = time.time()
    print(end - start)
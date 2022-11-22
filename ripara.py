import time
import openpyxl
from datetime import datetime
import os
import pandas as pd
import math

dataframe = pd.DataFrame(
    columns=['timestamp', 'carbon_intensity', 'low_emissions', 'renewable_emissions', 'total_production',
             'total_emissions'
        , 'nucleare_installed_capacity', 'nucleare_production', 'nucleare_emissions',
             'geotermico_installed_capacity', 'geotermico_production', 'geotermico_emissions'
        , 'biomassa_installed_capacity', 'biomassa_production', 'biomassa_emissions', 'carbone_installed_capacity',
             'carbone_production', 'carbone_emissions'
        , 'eolico_installed_capacity', 'eolico_production', 'eolico_emissions', 'fotovoltaico_installed_capacity',
             'fotovoltaico_production', 'fotovoltaico_emissions'
        , 'idroelettrico_installed_capacity', 'idroelettrico_production', 'idroelettrico_emissions',
             'accumuloidro_installed_capacity', 'accumuloidro_production', 'accumuloidro_emissions'
        , 'batterieaccu_installed_capacity', 'batterieaccu_production', 'batterieaccu_emissions',
             'gas_installed_capacity', 'gas_production', 'gas_emissions'
        , 'petrolio_installed_capacity', 'petrolio_production', 'petrolio_emissions',
             'sconosciuto_installed_capacity', 'sconosciuto_production', 'sconosciuto_emissions'
        , 'exchange_export', 'exchange_import'])

stati = ['El Hierro (Spagna)', 'Svezia', 'Francia', 'Danimarca orientale (Danimarca)', 'Romania',
         'Spagna', 'Belgio', 'Lettonia', 'Portogallo', 'Gran Canaria (Spagna)',
         'Danimarca occidentale (Danimarca)', 'Ungheria', 'Formentera (Spagna)', 'Lussemburgo', 'Finlandia',
         'Austria', 'Slovenia', 'Maiorca (Spagna)', 'Paesi Bassi', 'Lituania',
         'Slovacchia', 'Croazia', 'Ibiza (Spagna)', 'Polonia', 'Irlanda',
         'Sicilia (Italia)', 'Bulgaria', 'Tenerife (Spagna)', 'Germania', 'Centronord (Italia)',
         'Cipro', 'Minorca (Spagna)', 'Settentrione (Italia)', 'Estonia', 'Fuerteventura Lanzarote (Spagna)',
         'La Palma (Spagna)', 'Meridione (Italia)', 'La Gomera (Spagna)', 'Centrosud (Italia)', 'Sardegna (Italia)']

STATES_DIR = "statesRiparati"

def scarto(time):

    ora = int(time.split("-")[0].split(":")[0])
    scarto = 13 - ora
    minuti=0

    if scarto <= 0:
        minuti= int(time.split(" ")[0].split(":")[1])

        scarto = abs(scarto)
    elif scarto > 0:
        minuti= 50-int(time.split(" ")[0].split(":")[1])
        scarto = scarto - 1

    res=scarto*60+minuti
    x=2**(res/40)
    if x > 100:
        return 100
    else:
        return x


def buchi(lista) :
    x = 0
    c = 0
    l = []
    while (x < len(lista)) :
        cont = 0
        for i in lista[x] :
            if (i == None) :
                cont += 1
        if (cont >= 40) :
            c += 1
        else :
            cont = 0
        if (c > 0 and cont == 0) :
            l.append([c, x - c])
            c = 0
        x = x + 1

    print(l)
    return l


def riparaBuco(l, nuovofile) :
    file = []
    x = 0
    for i in l :
        buco = i[1]
        while (x < buco) :
            file.append(nuovofile[x])
            x += 1
        if (i[0] == 1) :
            tmp = []
            tmp.append(nuovofile[x][0])
            if (x <= 1) :
                y = 1
            else :
                y = -1
            j = 1
            while (j < len(nuovofile[x + y])) :
                tmp.append(nuovofile[x + y][j])
                j += 1

            file.append(tmp)
            x += 1

    while (x < len(nuovofile)) :
        file.append(nuovofile[x])
        x += 1
    return file


def riparaBuchigrandi(l, nuovofile) :
    file = []
    x = 0
    for i in l :
        buco = i[1]
        while (x < buco) :
            file.append(nuovofile[x])
            x += 1

        if (i[0] >= 15) :
            c = 0
            while (c < i[0]) :
                c += 1
                tmp = []
                tmp.append(nuovofile[x][0])
                if (x <= 144 + 1) :
                    y = 144
                else :
                    y = -144
                j = 1
                while (j < len(nuovofile[x + y])) :
                    tmp.append(nuovofile[x + y][j])
                    j += 1

                file.append(tmp)
                x += 1

    while (x < len(nuovofile)) :
        file.append(nuovofile[x])
        x += 1
    return file

def riparaBuchim(l, nuovofile) :
    file = []
    x = 0
    for i in l :
        buco = i[1]
        while (x < buco) :
            file.append(nuovofile[x])
            x += 1
        if (i[0] > 1 and i[0]<15) :
            tmp = []
            tmp.append(nuovofile[x][0])
            if (x <= 1) :
                y = 1
            else :
                y = -1
            j = 1
            while (j < len(nuovofile[x + y])) :
                tmp.append(nuovofile[x + y][j])
                j += 1
            #print("Primo ", tmp)
            file.append(tmp)
            x += 1
            c=2
            while (c < i[0]) :
                c+=1
                file.append(nuovofile[x])
                #print("xxxxx ", nuovofile[x])
                x += 1

            tmp = []
            tmp.append(nuovofile[x][0])
            if (x >= len(nuovofile)) :
                y = -1
            else :
                y = +1
            j = 1
            while (j < len(nuovofile[x + y])) :
                tmp.append(nuovofile[x + y][j])
                j += 1
            #print("ULTIMO ",tmp)
            file.append(tmp)
            x += 1

    while (x < len(nuovofile)) :
        file.append(nuovofile[x])
        x += 1
    return file


'''
12:10
12:40 -50%-2^-10 
13:00 -50%
13:10 
13:50
'''
def scala(time,value):
    '''
    13    50
    12:40 13:20 51,2
    12:30 13:30 51,4
    12:20 13:40 51,78
    12:10 13,50 52,15
    12 14 53,17
    11 15 66
    10 16 74
    9  17 82
    8  18 90
    7  19 100%
    '''

    #print(time,value)
    ora=int(time.split("-")[0].split(":")[0])
    #scarto = abs(13 - ora)
    #print(ora)
    if value!= None:

        if(ora>=19 or ora<=7):
            value=0
        else:
            value -= value * 50 / 100

            value -= value * (scarto(time)) / 100

    return value
if __name__ == '__main__' :
    statii = ['Germania']
    start = time.time()
    for s in stati:
        try :
            print(s)
            file_excel = openpyxl.load_workbook('states/' + s + '.xlsx')
            sheet_obj = file_excel.active
            numr = sheet_obj.max_row + 1
            numc = sheet_obj.max_column + 1
            print("numero di rig =", numr - 1)
            print("numero di col =", numc - 1)

            # ripara timestamp non alliniati a 00 10 20 30 40 50
            rnew = []

            c = []

            for j in range(1, numc) :
                c.append(sheet_obj.cell(row=1, column=j).value)
            rnew.append(c)

            for i in range(2, numr) :
                c = []
                for j in range(1, numc) :
                    tmp = sheet_obj.cell(row=i, column=j).value

                    if j != 1 :
                        c.append(tmp)

                    else :

                        if (tmp[4] != "0") :
                            print("#############################################################")
                            c.append(tmp[0 :4] + "0" + tmp[5 :])
                        else:
                            #print(tmp)
                            ''' 
                            riparazione mesi
                            datanew=tmp.split("-")
                            #print(datanew)
                            if int(datanew[1])+5<10:
                                datanew[1] = "0"+str(int(datanew[1])+5)
                            else:
                                datanew[1] = str(int(datanew[1]) + 5)
                            datanew="-".join(datanew)
                            #print(datanew)
                            tmp=datanew
                            '''
                            c.append(tmp)
                rnew.append(c)


            file_excel.close()
            #######################################################################################################################################################################

            # riparare i buchi causati dal crash del sistema
            starttime = timestamp = int(round(datetime.strptime(rnew[1][0], '%H:%M %d-%m-%Y').timestamp()))

            nuovofile = []
            nuovofile.append(rnew[0])

            for i in range(1, numr - 1) :
                timestamp = int(round(datetime.strptime(rnew[i][0], '%H:%M %d-%m-%Y').timestamp()))


                if (timestamp == starttime) :
                    nuovofile.append(rnew[i])

                else :
                    while (timestamp > starttime) :
                        timestampnew = datetime.fromtimestamp(starttime).strftime('%H:%M %d-%m-%Y')
                        nuovofile.append(
                            [timestampnew, None, None, None, None, None, None, None, None, None, None, None, None, None,
                             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])
                        starttime += 600

                    nuovofile.append(rnew[i])
                starttime += 600
            #######################################################################################################################################################################
            print(len(nuovofile))

            l = buchi(nuovofile)
            nuovofile = riparaBuco(l, nuovofile)
            print("Eliminati buchi pari a 1")

            l = buchi(nuovofile)
            rimedia=False
            for i in l :
                if (i[0] > 15) :
                    rimedia = True
                    break
            while(rimedia):
                nuovofile = riparaBuchigrandi(l, nuovofile)
                print("Eliminati buchi >= 15")
                l = buchi(nuovofile)
                rimedia=False
                for i in l:
                    if(i[0]>15):
                        rimedia=True
                        break


            rimedia = False
            for i in l :
                if (i[0] > 1 and i[0] < 15) :
                    rimedia = True
                    break
            while (rimedia):
                nuovofile = riparaBuchim(l, nuovofile)
                print("Eliminati buchi tra 2 e 14 inclusi")
                l = buchi(nuovofile)
                rimedia = False
                for i in l :
                    if (i[0] > 1 and i[0]<15) :
                        rimedia = True
                        break
            if(l!=[]):
                nuovofile = riparaBuco(l, nuovofile)
                print("Eliminati buchi pari a 1")
                l = buchi(nuovofile)

            print(len(nuovofile))
            #######################################################################################################################################################################
            # qui sotto gestisco il salvataggio del nuovo file
            #s = s + "NEW"

            path = os.path.join(STATES_DIR, s + ".xlsx")

            dfaaa = pd.DataFrame(columns=dataframe.columns)  # 44 colonne
            #print(nuovofile[0][22])
            #print(nuovofile[0][23])
            #print(nuovofile[1])
            for i in range(1, len(nuovofile)) :
                j = 0
                tmp = {}
                for df in dataframe :
                    if(isinstance(nuovofile[i][j], str)):
                        tmp[df] = nuovofile[i][j].replace(";","@")
                    else:

                        if (j==22 or j==23):
                            #print(nuovofile[1][j])
                            nuovofile[i][j]=scala(nuovofile[i][0],nuovofile[i][j])
                            tmp[df] = nuovofile[i][j]
                        else:
                            tmp[df] = nuovofile[i][j]
                    j += 1
                dfbbb = pd.concat([dataframe.copy(), pd.DataFrame(tmp, index=[0])], ignore_index=True)
                dfaaa = pd.concat([dfaaa, dfbbb])

            try :
                if (len(s) <= 30) :

                    dfaaa.to_excel(path, sheet_name=s, index=False)
                else :
                    s1 = s.split(" ")[0]
                    dfaaa.to_excel(path, sheet_name=s1, index=False)

            except Exception as e :
                pass

        except Exception as e :
            print(e)

    end = time.time()
    print(end - start)
